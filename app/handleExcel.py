from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import colors, fills, Font, Border, Side
import config as c
import re
import os

""" 
# cell = worksheet.cell(number, letter (number))
# Cells B7 - B26 = Date
# Cells C7 - C26 = Description
# Cells D7 - D26 = Miles
# Cells E7 - E26 = Account code
# Cells F7 - F26 = Amount
# info = [date, amount of miles and description, amount of miles, account code, total calculated]
"""


def requirements(name: list, date, book_name):  # Add dates and names
    """
    :param name: user name [first, last]
    :param date: date
    :param book_name: book file name
    :return: file
    """
    workbook = get_book(book_name)
    worksheet = workbook["Expense Claim Form 14-11-19"]
    bottom_row = 28
    if find_available_row(worksheet) >= 27:
        bottom_row = find_available_row(worksheet) + 1  # avoid overflow
    # Name at top
    for i in range(2, 4):
        name_cell = worksheet.cell(4, i)
        name_cell.value = name[i - 2]
    # Date at bottom
    date_cell = worksheet.cell(bottom_row, 6)
    date_cell.value = date
    # Name at bottom
    name_cell = worksheet.cell(bottom_row, 3)
    name_cell.value = name[0] + " " + name[1]
    workbook.save(c.Config.RECLAIM_ROUTE + book_name)


def edit_row(info: list, book_name, row=None):  # Edit a row
    """
    :param info: info = [date, amount of miles and description, amount of miles, account code, total calculated]
    :param book_name: file name of sheet
    :param row: row number to edit
    """
    workbook = get_book(book_name)  # Get workbook
    worksheet = workbook["Expense Claim Form 14-11-19"]  # Open worksheet
    if not row:
        # If no explicit row is given
        row = find_available_row(worksheet)  # Find the next available row
        if worksheet.cell(row, 5).value:
            # If there is no free space
            merged_cells_range = worksheet.merged_cells.ranges
            # Move all unmerged cells
            for merged_cell in merged_cells_range:
                if (
                    ord(re.sub(r"\d+", "", str(merged_cell).split(":")[0]).lower()) - 96
                    < 7
                ):  # If cell letter is > F
                    merged_cell.shift(0, 1)  # Move unmerged cell down by one
            worksheet.insert_roworksheet(row)  # Insert a new row
    cell = worksheet.cell(row, 1)
    cell.value = row - 6  # Add row index
    for column in range(2, 7):  # column B - column F
        # Fill in data
        cell = worksheet.cell(row, column)
        cell.value = info[column - 2]
        # Consistent formatting
        grey = colors.Color(rgb="D9D9D9")
        cell.fill = fills.PatternFill(patternType="solid", fgColor=grey)
        if column == 6:  # Column F
            # Excel formatting as currency (source: Excel)
            cell.number_format = '_-[$£-en-GB]* #,##0.00_-;-[$£-en-GB]* #,##0.00_-;_-[$£-en-GB]* "-"??_-;_-@_-'
        cell.font = Font(bold=False)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell.border = thin_border  # Persistent border style
    workbook.save(c.Config.RECLAIM_ROUTE + book_name)  # Save reclaim form


def add_images(book_name, row, filename):
    """
    Add images to excel file
    :param book_name: file name of form
    :param row: row number
    :param filename: image filename
    """
    workbook = get_book(book_name)  # Get workbook
    sheet_name = "Receipt for row " + str(row)
    workbook.create_sheet(sheet_name)  # Create a new sheet
    worksheet = workbook[sheet_name]
    img = Image(c.Config.IMAGE_UPLOADS + filename)
    img.anchor = "A1"
    worksheet.add_image(img)  # Add image
    workbook.save(c.Config.RECLAIM_ROUTE + book_name)  # Save workbook


def add_signature(signature, book_name, date):
    """
    Add a signature to excel sheet
    :param signature: signature file name
    :param book_name: file name of reclaim form
    :param date: date signed
    """
    workbook = get_book(book_name)
    worksheet = workbook["Expense Claim Form 14-11-19"]
    img = Image(c.Config.SIGNATURE_ROUTE + signature)
    img.anchor = "C29"
    worksheet.add_image(img)
    cell = worksheet.cell(29, 6)
    cell.value = date
    workbook.save(c.Config.RECLAIM_ROUTE + book_name)


def find_available_row(worksheet):
    """
    Find next empty row
    :param worksheet: worksheet object
    :return: int row
    """
    row = 7
    while worksheet.cell(row, 2).value:
        row += 1
    return row


def delete_all_sheets():
    """
    deletes all files in reclaim folder (to clear space)
    """
    files = [f for f in os.listdir(c.Config.RECLAIM_ROUTE)]
    for f in files:
        os.remove(os.path.join(c.Config.RECLAIM_ROUTE, f))


def create_new_sheet(book_name):
    """
    create a new excel sheet
    :param book_name: file name of book
    """
    workbook = load_workbook(c.Config.STATIC + "Expenses form.xlsx")
    workbook.save(c.Config.RECLAIM_ROUTE + book_name)


def get_book(book_name):
    """
    :param book_name: filename of expenses form
    """
    try:
        load_workbook(c.Config.RECLAIM_ROUTE + book_name)
    except FileNotFoundError:
        create_new_sheet(book_name)
    return load_workbook(c.Config.RECLAIM_ROUTE + book_name)

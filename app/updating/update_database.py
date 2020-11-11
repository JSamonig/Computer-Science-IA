from openpyxl import load_workbook
from app.models import Account_codes, cost_centres
from app import db


def first():  # Adds account codes from excel list
    wb = load_workbook("departments.xlsx")
    ws = wb["Sheet1"]
    for i in range(2, 250):
        myarr = []
        for j in range(1, 3):
            myarr.append(ws.cell(i, j).value)
            print(type(ws.cell(i, j).value))
            if ws.cell(i, j).value is None:
                db.session.commit()
                print("done")
                return i
        account = Account_codes(account_id=myarr[0], account_name=myarr[1])  # add to account codes
        db.session.add(account)


def second():  # adds cost centres from excel list
    wb = load_workbook("names.xlsx")
    ws = wb["Sheet1"]
    for i in range(2, 2200):
        myarr = []
        for j in range(1, 5):
            myarr.append(ws.cell(i, j).value)
            print(ws.cell(i, j).value)
            if ws.cell(i, j).value is None:
                db.session.commit()
                return i
        account = db.session.query(Account_codes).filter_by(account_id=myarr[2]).first()
        if account is not None:
            account.cost_centre = myarr[1]
        new_centre = cost_centres(cost_centre_id=myarr[2], purpose_cost_centre=myarr[3], purpose_id=myarr[0])
        db.session.add(new_centre)  # add cost centres


def main():
    first()
    second()


main()  # run to initialize database
